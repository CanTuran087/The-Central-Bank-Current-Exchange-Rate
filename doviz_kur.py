import requests
import xml.etree.ElementTree as ET
import openpyxl
import tkinter as tk
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import messagebox
from datetime import datetime


def getExchangeRates():
    bugunTarih = datetime.now().strftime("%d-%m-%Y")
    apiKey = '' #evds2.tcmb.gov.tr sitesinden aldığınız API Key'i ekleyip çalıştırabilirsiniz.
    url = 'https://evds2.tcmb.gov.tr/service/evds/series=TP.DK.USD.A-TP.DK.EUR.A-TP.DK.CHF.A-TP.DK.GBP.A-TP.DK.JPY.A-TP.DK.USD.S-TP.DK.EUR.S-TP.DK.CHF.S-TP.DK.GBP.S-TP.DK.JPY.S&startDate=' + \
        bugunTarih + '&endDate=' + bugunTarih + '&type=xml&key=' + apiKey + ''

    # API den XML verisini al
    print(url)
    response = requests.get(url=url)
    xmlData = response.text

    # XML verisini parse et
    root = ET.fromstring(xmlData)

    # <TP_DK_USD_A> etiketindeki değeri al
    # Verileri sözlükte topla
    exhangeRates = {
        'usdAlis': root.find('.//TP_DK_USD_A').text,
        'usdSatis': root.find('.//TP_DK_USD_S').text,
        'eurAlis': root.find('.//TP_DK_EUR_A').text,
        'eurSatis': root.find('.//TP_DK_EUR_S').text,
        'chfAlis': root.find('.//TP_DK_CHF_A').text,
        'chfSatis': root.find('.//TP_DK_CHF_S').text,
        'gbpAlis': root.find('.//TP_DK_GBP_A').text,
        'gbpSatis': root.find('.//TP_DK_GBP_S').text,
        'jpyAlis': root.find('.//TP_DK_JPY_A').text,
        'jpySatis': root.find('.//TP_DK_JPY_S').text
    }
    # print(f'Dolar Alış Kuru: {usdAlis} Dolar Satış Kuru: {usdSatis}')

    return exhangeRates


def pathControl(filePath):
    if os.path.exists(filePath):
        root = tk.Tk()
        root.withdraw()  # Ana pencereyi gizle

        cevap = messagebox.askquestion(
            'Uyarı', 'Belirtilen dosya zaten var, üzerine yazmak istiyor musunuz ?')

        if cevap == 'yes':
            isActive = 2

        else:
            messagebox.showinfo("Uyarı", "Üzerine yazma işlemi iptal edildi.")
            isActive = 0

    else:
        isActive = 1

    return isActive


def main():
    exhangeRates = getExchangeRates()

    # Kullanıcının masaüstünü bulur ve excel pathini ekler
    desktopPath = os.path.join(os.path.expanduser('~'), 'Desktop')
    filePath = os.path.join(desktopPath, 'kurlar_excel.xlsx')

    # Hücreyi sarı renk ve kalın yapar
    yellowColor = PatternFill(start_color='FFFF00',
                              end_color='FFFF00', fill_type='solid')
    boldFont = Font(bold=True)

    # Kenarlık
    sideStyle = Side(style='thin')
    thinBorder = Border(left=sideStyle, right=sideStyle, top=sideStyle, bottom=sideStyle)

    isActive = pathControl(filePath)

    if isActive == 1:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Kurlar'

    elif isActive == 2:
        workbook = openpyxl.load_workbook(filePath)
        sheet = workbook.active
        sheet.title = 'Kurlar'

    else:
        exit()

    sheet['B1'] = 'Döviz Alış'
    sheet['C1'] = 'Döviz Satış'

    sheet['B1'].fill = yellowColor
    sheet['C1'].fill = yellowColor
    sheet['B1'].font = boldFont
    sheet['C1'].font = boldFont
    sheet['B1'].border = thinBorder
    sheet['C1'].border = thinBorder

    sheet['A2'] = 'Dolar'
    sheet['A3'] = 'Euro'
    sheet['A4'] = 'İsviçre Frangı'
    sheet['A5'] = 'İngiliz Sterlini'
    sheet['A6'] = 'Japon Yeni'

    # Verileri hücre içine al
    row = 2 
    row2 = 2
    for cellValue in ['usdAlis', 'eurAlis', 'chfAlis', 'gbpAlis', 'jpyAlis','usdSatis', 'eurSatis', 'chfSatis', 'gbpSatis', 'jpySatis']:

        if row < 7:
            sheet[f'B{row}'] = exhangeRates[cellValue]
            sheet[f'B{row}'].border = thinBorder

        elif row >= 7:
            sheet[f'C{row2}'] = exhangeRates[cellValue]
            sheet[f'C{row2}'].border = thinBorder
            row2 += 1

        row += 1

    row = 2
    while row <= 6:
        sheet[f'A{row}'].fill = yellowColor
        sheet[f'A{row}'].font = boldFont
        sheet[f'A{row}'].border = thinBorder
        row += 1

    workbook.save(filePath)

    messagebox.showinfo("Bilgi", "Excel dosyası başarıyla kaydedildi.")


if __name__ == '__main__':
    main()
